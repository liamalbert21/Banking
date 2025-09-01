import re, hashlib, openpyxl, bank_data
from openpyxl.utils.exceptions import InvalidFileException

def login(pin = None, admin_login = False):
    if pin:
        verify = pin; input_type = int
    else:
        verify = bank_data.Bank.get_password(); input_type = str
    for x in range(3):
        attempt = get_input(f'Attempt {x+1}: ', input_type)
        if not pin:
            if hashlib.sha512(attempt.encode()).digest() == verify:
                return True
        elif attempt == verify:
            return True
        print('Incorrect')
    print('Login failed. User lockout') if admin_login else print('Login failed.\n')
    return False

def get_input(prompt, cast):
    while True:
        try:
            return cast(input(prompt))
        except ValueError:
            print('Error: Invalid input')

def verify_account(n = 1, confirm_number = False, need_number = False):
    def decorator(func):
        def wrapper(*args):
            send = []
            for i in range(n):
                account_type = ' ' if n == 1 else ' origin ' if n == 2 and i == 0 else ' destination '
                account_number = get_input(f'Enter{account_type}account number: ', int)
                account = bank_data.Bank.get_account(account_number)
                if not account:
                    return
                elif account.is_locked:
                    print('Error: Account is locked\n')
                    return
                elif account.pin:
                    print('Enter account pin. You have three attempts')
                    if not login(pin = account.pin):
                        return
                if confirm_number:
                    if account_number != get_input('Re-enter account number: ', int):
                        print('Error: Accounts do not match\n')
                        return
                if need_number:
                    send.append(account_number)
                else:
                    send.append(account_number)
            func(*args, *send)
        return wrapper
    return decorator

def new_account():
    name = get_input("Enter account holder's name: ", str)
    valid_name = True if not re.search(r'[^a-zA-Z\s]', name) else False
    if not valid_name:
        print("Error: Name must only contain alphabetic characters and whitespace\n")
        return
    start = get_input('Enter initial deposit: $', float)
    bank_data.Bank.create_account(name, start)

@verify_account()
def add_remove_pin(account):
    if account.pin:
        account.pin = None
    else:
        pin = get_input('Enter desired pin: ', int)
        verify = get_input('Re-enter desired pin: ', int)
        if verify == pin:
            account.pin = pin
        else:
            print('Error: Pins do not match\n')
            return
    print('Success.\n')

@verify_account()
def change_funds(is_deposit, account):
    action = 'deposit' if is_deposit else 'withdrawal'
    amount = get_input(f'Enter {action} amount: $', float)
    if is_deposit:
        account.deposit(amount)
    else:
        account.withdraw(amount)

@verify_account(n = 2)
def move_funds(origin, destination):
    amount = get_input('Enter transfer amount: $', float)
    result = origin.withdraw(amount, True)
    if result:
        destination.deposit(amount, True)

@verify_account()
def view(is_balance, account):
    if is_balance:
        account.check_balance()
    else:
        account.print_statement()

@verify_account(confirm_number = True, need_number = True)
def discard_account(account_number):
    print('Warning: This action is irreversible')
    print('Enter admin password to confirm deletion. You have 3 attempts')
    if login():
        bank_data.Bank.remove_account(account_number)

@verify_account(confirm_number = True)
def lock_account(account):
    access = 'locked' if account.is_locked else 'unlocked'
    print(f'This account is {access}')
    print('Enter admin password to change account status. You have 3 attempts')
    if login():
        account.is_locked = not account.is_locked
        print('Success.\n')

def import_export(is_import):
    print('Enter admin password. You have 3 attempts')
    if login():
        if is_import:
            path = get_input('Enter file path: ', str)
            try:
                worksheet = openpyxl.load_workbook(path.strip(' "'), data_only = True).active
                print('In progress...')
                max_row = worksheet.max_row
                for row in range(2, max_row + 1):
                    name = worksheet.cell(row, 1).value
                    start = worksheet.cell(row, 2).value
                    status = True if worksheet.cell(row, 4).value else False
                    bank_data.Bank.create_account(name, start, status, True)
                print('Success.\n')
            except FileNotFoundError:
                print('Error: Invalid path\n')
            except InvalidFileException:
                print('Error: Invalid file format\n')
        else:
            bank_data.Bank.export_database()

def new_password():
    print('Enter current admin password. You have 3 attempts')
    if login():
        desired_password = get_input('Enter new password: ', str)
        verification = get_input('Re-enter new password: ', str)
        if desired_password == verification:
            bank_data.Bank.change_password(desired_password)
        else:
            print('Error: Passwords do not match\n')