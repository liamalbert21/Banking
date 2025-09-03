import random, hashlib, datetime, openpyxl
from openpyxl.styles import Font, Alignment

class BankAccount:
    def __init__(self, holder, balance, is_locked):
        self.holder = holder
        self.balance = balance
        self.transactions = [{
            'Type': 'Initialization',
            'Amount': balance,
            'Balance': balance,
            'Timestamp': datetime.datetime.now()
        }]
        self.is_locked = is_locked
        self.pin = None

    def deposit(self, amount, is_transfer = False):
        if amount <= 0:
            print('Error: Deposit amount must be positive\n')
        else:
            self.balance += amount
            self.transactions.append({
                'Type': 'Deposit',
                'Amount': amount,
                'Balance': self.balance,
                'Timestamp': datetime.datetime.now()
            })
            print(f'Success. New balance: ${self.balance:.2f}\n') if not is_transfer else print('Success.\n')

    def withdraw(self, amount, is_transfer = False):
        action = 'Transfer' if is_transfer else 'Withdrawal'
        if amount <= 0:
            print(f'Error: {action} amount must be positive\n')
            return False
        elif amount > self.balance:
            print(f'Error: {action} amount cannot exceed balance\n')
            return False
        else:
            self.balance -= amount
            self.transactions.append({
                'Type': 'Withdrawal',
                'Amount': amount,
                'Balance': self.balance,
                'Timestamp': datetime.datetime.now()
            })
            if not is_transfer:
                print(f'Success. New balance: ${self.balance:.2f}\n')
        return True

    def check_balance(self):
        print(f'Current balance: ${self.balance:.2f}\n')

    def print_statement(self):
        print(f'\n=== {self.holder} ===')
        for i, x in enumerate(self.transactions):
            print(f'Transaction {i+1}:')
            for k, v in x.items():
                if k == 'Type':
                    print(f'- {k}: {v}')
                elif k in ('Amount', 'Balance'):
                    print(f'- {k}: ${v:.2f}')
                else:
                    print(f'- {k}: {v.strftime("%Y-%m-%d %H:%M:%S")}')
        print()

class Bank:
    __stored_accounts = {}
    __available_accounts = set(range(1000000, 10000000))
    __admin_password = hashlib.sha512('password'.encode()).digest()

    @classmethod
    def create_account(cls, holder, initial_deposit, is_locked = False, is_import = False):
        if initial_deposit >= 0:
            account_number = random.sample(list(cls.__available_accounts), 1)[0]
            cls.__available_accounts.remove(account_number)
            cls.__stored_accounts[account_number] = BankAccount(holder, initial_deposit, is_locked)
            if not is_import:
                print(f'Success. Account number: {account_number}\n')
        else:
            print('Error: Initial balance must be nonnegative\n')

    @classmethod
    def get_account(cls, account_number):
        try:
            return cls.__stored_accounts[account_number]
        except KeyError:
            print('Error: No such account exists\n')

    @classmethod
    def get_password(cls):
        return cls.__admin_password

    @classmethod
    def remove_account(cls, account_number):
        cls.__stored_accounts.pop(account_number)
        cls.__available_accounts.add(account_number)
        print('Success.\n')

    @classmethod
    def change_password(cls, password):
        cls.__admin_password = hashlib.sha512(password.encode()).digest()
        print('Success.\n')

    @classmethod
    def export_database(cls):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Database'
        headers = ('Account Number', 'Account Name', 'Balance', 'Status', 'Status (B)')
        for i in range(len(headers)):
            cell = worksheet.cell(1, i + 1, headers[i])
            cell.font = Font(bold = True)
            cell.alignment = Alignment(horizontal = 'center')
        for row, (account_number, account) in enumerate(cls.__stored_accounts.items()):
            if account.is_locked:
                status_b = 1; status = 'Locked'
            else:
                status_b = 0; status = 'Unlocked'
            data = (account_number, account.holder, account.balance, status, status_b)
            for i in range(len(data)):
                worksheet.cell(row + 2, i + 1, data[i])
        workbook.save('Database.xlsx')
        print('Success.\n')